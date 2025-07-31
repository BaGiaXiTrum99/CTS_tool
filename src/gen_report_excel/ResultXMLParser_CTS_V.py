import os
import math
import logging
import xml.etree.cElementTree as ET

from xml.etree.ElementTree import Element
from datetime import datetime
from openpyxl import Workbook
from utils.constants import *
from utils.time_caculation import *
from utils.ExcelHandler import *

logger = logging.getLogger("cts_logger." + __name__)

class ResultXMLParser_CTS_V:
    def __init__(self, path: str, time_unit : str) -> None:
        logger.info("Generate Report CTS Verifier Feature")
        self.result_path = f"{path}/test_result.xml" 
        self.time_unit = time_unit
        self.__root = self.__read_result_file()
    
    def __read_result_file(self):
        logger.info(f"Reading XML result file {self.result_path}")
        tree = ET.parse(self.result_path)   
        return tree.getroot()     
    
    def __write_module_row(self, ws: Worksheet, index: int, module_data: dict):
        row = [index] + [module_data[col.value] for col in list(ReportColumnCTS_V)[1:]]
        ws.append(row)

    def get_module_runner(self):
        list_module = self.__root.findall('.Module')
        logger.info(f"Found {len(list_module)} module(s) in result file {self.result_path}")
        assert len(list_module) == 1, "There are more then 1 module in CTS Verifier report"
        return list_module[0]

    def parser_all_modules(self):
        wb = Workbook()

        meta_ws = wb.create_sheet(title="Metadata",index = 0)
        ExcelHandler.write_metadata_sheet(meta_ws,self.__root)

        ws : Worksheet = wb.create_sheet(title="CTS-V Report",index = 1)
        ws.column_dimensions['B'].width = 70
        # Ghi header
        headers = [col.value for col in ReportColumnCTS_V]
        ws.append(headers)
        totals = {
            ReportColumn.EXECUTION_TIME.value: "0s" if self.time_unit == TimeUnit.HMS.value else 0
        }
        module = self.get_module_runner()
        module_name = module.get('name')
        assert module_name is not None, "Can not get name of module"

        tests = module.findall(".//TestCase/Test")
        logger.info(f"Found {len(tests)} test(s) in module, start parsing")
        idx = 0
        for test in tests:
            test_infor = self.__parser_test(test)
            idx += 1
            self.__write_module_row(ws, idx, test_infor)
            if self.time_unit == TimeUnit.HMS.value:
                totals[ReportColumn.EXECUTION_TIME.value] = sum_durations(totals[ReportColumn.EXECUTION_TIME.value],
                                                                            test_infor[ReportColumn.EXECUTION_TIME.value])
            else:
                totals[ReportColumn.EXECUTION_TIME.value] += sum(
                    int(x) for x in re.findall(r'\d+', test_infor[ReportColumn.EXECUTION_TIME.value])
                )     
        if self.time_unit == TimeUnit.S.value:
            totals[ReportColumn.EXECUTION_TIME.value] = str(totals[ReportColumn.EXECUTION_TIME.value]) + "s" 
        elif self.time_unit == TimeUnit.MS.value:
            totals[ReportColumn.EXECUTION_TIME.value] = str(totals[ReportColumn.EXECUTION_TIME.value]) + "ms" 
        ws.column_dimensions['D'].width = 15
        total_row = ["Total", "", "", totals[ReportColumn.EXECUTION_TIME.value]]
        ws.append(total_row)

        output_file = f"./result/myresult_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        wb.save(output_file)
        logger.info(f"Saved result to {output_file}")

    def __parser_test(self, test : Element) -> dict:
        name = test.get("name")
        logger.info(f"Parsing test {name}")

        result = test.get("result")

        start_end_times = test.findall(".//RunHistory/Run")
        start_time_list = []
        end_time_list = []
        for start_end_time in start_end_times:
            start_time_list.append(int(start_end_time.get("start")))
            end_time_list.append(int(start_end_time.get("end")))
        start_time = min(start_time_list)    
        end_time = max(end_time_list)
        execution_time = end_time - start_time
        if self.time_unit == TimeUnit.S.value:
            execution_time = str(math.ceil(execution_time/1000))
        elif self.time_unit == TimeUnit.HMS.value:
            execution_time = format_duration(math.ceil(execution_time/1000))

        return {
            ReportColumnCTS_V.TEST.value     : name,
            ReportColumnCTS_V.RESULT.value   : result,
            ReportColumnCTS_V.EXECUTION_TIME.value: execution_time
        }


    
