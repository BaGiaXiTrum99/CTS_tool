import os
import re
import logging
import xml.etree.cElementTree as ET
from xml.etree.ElementTree import Element,ElementTree
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime

from utils.time_caculation import TimeHandler
from utils.constants import *
from utils.ExcelHandler import ExcelHandler

logger = logging.getLogger("cts_logger." + __name__)

class MultiResultXMLParser:
    def __init__(self, path: str, time_unit : str, output_dir: str ) -> None:
        logger.info("Generate Multi Report CTS Feature")
        self.result_path = os.path.abspath(path)
        self.time_unit = time_unit
        self.output_dir = output_dir

    def __search_all_single_result_folder(self):
        report_files = []
        for subfolder in os.scandir(self.result_path):
            if subfolder.is_dir():
                if os.path.exists(subfolder.path+'/test_result.xml'):
                    logger.info(f'Found valid folder path {subfolder.path}')
                    report_files.append(subfolder.path+'/test_result.xml')
                else:
                    logger.error(f'Folder {subfolder.path} is not containing test_result.xml, can not parse')
        assert len(report_files), 'Length of results not be zero!!!'
        return sorted(report_files)

    def __get_root_of_result_file(self,report_file):
        logger.info(f"Reading XML result file {report_file}")
        try:
            tree = ET.parse(report_file)
        except ET.ParseError as e:
            msg = f"Failed to parse XML file {report_file}: {e}"
            logger.error(msg)
            raise RuntimeError(msg)
        return tree.getroot()     

    def __get_list_modules_name(self, root : ElementTree):
        list_modules = root.findall('.Module')
        logger.info(f"Found {len(list_modules)} module(s)")
        return [module.get('name') for module in list_modules]
    
    def __parser_module_full(self,module : Element ) -> dict:
        name = module.get("name")
        logger.info(f"Parsing module {name}")

        execution_time = TimeHandler.get_execution_time_from_module(module,self.time_unit)
        
        assert module.get("done") is not None, "Can not get result of module"
        done = True if module.get("done") == "true" else False
        logger.debug(f"Done: {done}")

        test_cases = module.findall(".TestCase")
        logger.info(f"Found {len(test_cases)} test cases")

        pass_count = 0
        fail_count = 0
        assumption_failure = 0
        ignored = 0
        for test_case in test_cases:
            logger.debug(f"Parsing test case {test_case.get('name')}")
            for test_step in test_case.findall(".Test"):
                test_step_name = test_step.get('name')
                test_step_result = test_step.get("result")
                if test_step_result == "fail":
                    fail_count += 1
                    logger.debug(f"Test {test_step_name} failed")
                elif test_step_result == "ASSUMPTION_FAILURE":
                    assumption_failure += 1
                    logger.debug(f"Test {test_step_name} assumption failure")
                elif test_step_result == "IGNORED":
                    ignored += 1
                    logger.debug(f"Test {test_step_name} ignored")
                else:
                    pass_count += 1
                    logger.debug(f"Test {test_step_name} passed")

        total_tests_str = module.get("total_tests")
        assert total_tests_str is not None, "Can not get total test of module"
        logger.info(f"Found {total_tests_str} tests")
        total_testcases = int(total_tests_str) if total_tests_str else 0

        if total_testcases != pass_count + fail_count + assumption_failure + ignored:
            msg = f"Mismatch test count in {name}: total={total_testcases}, counted={pass_count + fail_count + assumption_failure + ignored}"
            logger.error(msg)
            raise ValueError(msg)
        
        return {
            ReportColumn.MODULES.value  : name,
            ReportColumn.PASSED.value   : pass_count,
            ReportColumn.FAILED.value   : fail_count,
            ReportColumn.ASSUMPTION_FAILURE.value : assumption_failure,
            ReportColumn.IGNORED.value  : ignored,
            ReportColumn.TOTAL.value    : total_testcases,
            ReportColumn.DONE.value     : done,
            ReportColumn.EXECUTION_TIME.value: execution_time
        }

    def __parser_module_time_execution(self,module : Element ) -> str:
        name = module.get("name")
        logger.info(f"Parsing module {name}")
        return TimeHandler.get_execution_time_from_module(module,self.time_unit)

    def __write_module_row(self, ws: Worksheet, index: int, module_data: dict):
        row = [index] + [module_data[col.value] for col in list(ReportColumn)[1:]]
        ws.append(row)

    def parsing_all_results_file(self):
        wb = Workbook()

        meta_ws = wb.create_sheet(title="Metadata",index = 0)

        ws : Worksheet = wb.create_sheet(title="CTS Report",index = 1)
        ws = ExcelHandler.create_header_row(ws)

        totals = {
            ReportColumn.PASSED.value: 0,
            ReportColumn.FAILED.value: 0,
            ReportColumn.ASSUMPTION_FAILURE.value: 0,
            ReportColumn.IGNORED.value: 0,
            ReportColumn.TOTAL.value: 0,
            ReportColumn.DONE.value: True,
            ReportColumn.EXECUTION_TIME.value: "0s" if self.time_unit == TimeUnit.HMS.value else 0
        }

        report_files = self.__search_all_single_result_folder()
        # Loại bỏ các phần tử phía trước, chỉ parse dữ liệu cụ thể từ phần tử cuối
        last_report_file = report_files.pop(-1)

        # Xu li file cuoi truoc
        last_root = self.__get_root_of_result_file(last_report_file)
        logger.info(f"Parsing XML result file {last_report_file}")
        ExcelHandler.write_metadata_sheet(meta_ws,last_root)

        list_module_name = self.__get_list_modules_name(last_root)
        idx_row = 1
        for module_name in list_module_name:
            logger.info(f"Found module {module_name} existed in DUT, start parsing")
            module = last_root.find(f'.//Module[@name="{module_name}"]')
            module_infor = self.__parser_module_full(module)

            old_module_execution_time = module_infor[ReportColumn.EXECUTION_TIME.value]
            for report_file in report_files:
                logger.info(f"Parsing XML result file {report_file}")
                root = self.__get_root_of_result_file(report_file)
                module = root.find(f'.//Module[@name="{module_name}"]')
                if module is None:
                    logger.error(f"Module {module_name} not found in {report_file}")
                    continue
                module_execution_time = self.__parser_module_time_execution(module)
                if old_module_execution_time != module_execution_time:
                    module_infor[ReportColumn.EXECUTION_TIME.value] = TimeHandler.sum_durations(module_infor[ReportColumn.EXECUTION_TIME.value],module_execution_time)
                    old_module_execution_time = module_execution_time
                else:
                    continue
                logger.debug(module_infor)
            self.__write_module_row(ws, idx_row , module_infor)
            idx_row += 1
            for key in (ReportColumn.PASSED.value, 
                        ReportColumn.FAILED.value, 
                        ReportColumn.ASSUMPTION_FAILURE.value, 
                        ReportColumn.IGNORED.value, 
                        ReportColumn.TOTAL.value):
                totals[key] += module_infor[key]
            if self.time_unit == TimeUnit.HMS.value:
                totals[ReportColumn.EXECUTION_TIME.value] = TimeHandler.sum_durations(totals[ReportColumn.EXECUTION_TIME.value],
                                                                        module_infor[ReportColumn.EXECUTION_TIME.value])
            else:
                totals[ReportColumn.EXECUTION_TIME.value] += sum(
                    int(x) for x in re.findall(r'\d+', module_infor[ReportColumn.EXECUTION_TIME.value])
                )                    
            totals[ReportColumn.DONE.value] = totals[ReportColumn.DONE.value] and module_infor[ReportColumn.DONE.value]

        if self.time_unit == TimeUnit.S.value:
            totals[ReportColumn.EXECUTION_TIME.value] = str(totals[ReportColumn.EXECUTION_TIME.value]) + "s" 
        elif self.time_unit == TimeUnit.MS.value:
            totals[ReportColumn.EXECUTION_TIME.value] = str(totals[ReportColumn.EXECUTION_TIME.value]) + "ms" 

        total_row = ["Total", ""] + [totals[k] for k in [col.value for col in ReportColumn][2:]]
        ws.append(total_row)

        output_file = f"{self.output_dir}/myresult_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        wb.save(output_file)
        logger.info(f"Saved result to {output_file}")
