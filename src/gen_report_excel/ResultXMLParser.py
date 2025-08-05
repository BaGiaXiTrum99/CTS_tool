import os
import re
import math
import logging
import xml.etree.cElementTree as ET

from xml.etree.ElementTree import Element
from datetime import datetime
from openpyxl import Workbook
from utils.constants import *
from utils.time_caculation import *
from utils.ExcelHandler import *
from src.gen_report_excel.SubplansParser import SubplansParser

logger = logging.getLogger("cts_logger." + __name__)

class ResultXMLParser:
    def __init__(self, path: str, time_unit : str) -> None:
        logger.info("Generate Report Feature")
        self.result_path = f"{path}/test_result.xml"
        self.time_unit = time_unit
        self.__root = self.__read_result_file()
    
    def __read_result_file(self):
        logger.info(f"Reading XML result file {self.result_path}")
        tree = ET.parse(self.result_path)   
        return tree.getroot()     
    
    def __write_module_row(self, ws: Worksheet, index: int, module_data: dict):
        row = [index] + [module_data[col.value] for col in list(ReportColumn)[1:]]
        ws.append(row)

    def __parse_module_and_fill_excel(self,module : Element, totals_row : dict, ws : Worksheet, idx_row : int):
        module_infor = self.__parser_module(module)
        for key in (ReportColumn.PASSED.value, 
                    ReportColumn.FAILED.value, 
                    ReportColumn.ASSUMPTION_FAILURE.value, 
                    ReportColumn.IGNORED.value, 
                    ReportColumn.TOTAL.value):
            totals_row[key] += module_infor[key]
        if self.time_unit == TimeUnit.HMS.value:
            totals_row[ReportColumn.EXECUTION_TIME.value] = sum_durations(totals_row[ReportColumn.EXECUTION_TIME.value],
                                                                    module_infor[ReportColumn.EXECUTION_TIME.value])
        else:
            print(totals_row[ReportColumn.EXECUTION_TIME.value])
            print(re.findall(r'\d+', module_infor[ReportColumn.EXECUTION_TIME.value]))
            totals_row[ReportColumn.EXECUTION_TIME.value] += sum(
                int(x) for x in re.findall(r'\d+', module_infor[ReportColumn.EXECUTION_TIME.value])
            )                    
        totals_row[ReportColumn.DONE.value] = totals_row[ReportColumn.DONE.value] and module_infor[ReportColumn.DONE.value]
        self.__write_module_row(ws, idx_row, module_infor)

    def get_module_runner(self):
        list_module = self.__root.findall('.Module')
        logger.info(f"Found {len(list_module)} module(s) in result file {self.result_path}")
        return list_module

    def parser_all_modules(self,subplans_path : str | None):
        wb = Workbook()

        meta_ws = wb.create_sheet(title="Metadata",index = 0)
        ExcelHandler.write_metadata_sheet(meta_ws,self.__root)

        ws : Worksheet = wb.create_sheet(title="CTS Report",index = 1)
        ws.column_dimensions['B'].width = 30
        # Ghi header
        headers = [col.value for col in ReportColumn]
        ws.append(headers)
        totals = {
            ReportColumn.PASSED.value: 0,
            ReportColumn.FAILED.value: 0,
            ReportColumn.ASSUMPTION_FAILURE.value: 0,
            ReportColumn.IGNORED.value: 0,
            ReportColumn.TOTAL.value: 0,
            ReportColumn.DONE.value: True,
            ReportColumn.EXECUTION_TIME.value: "0s" if self.time_unit == TimeUnit.HMS.value else 0
        }
        modules_runner = self.get_module_runner()
        if subplans_path is not None:
            subplan_parser = SubplansParser(subplans_path)
            module_list_from_subplans = subplan_parser.get_module_list_from_subplans()
            
            for idx in range(len(module_list_from_subplans)):
                module_name_from_subplans = module_list_from_subplans[idx]
                matched = False
                for module in modules_runner:
                    module_name = module.get('name')
                    assert module.get('name') is not None, "Can not get name of module"
                    if module_name_from_subplans == module_name:
                        matched = True
                        logger.info(f"Found module {module_name} existed in DUT, start parsing")
                        self.__parse_module_and_fill_excel(module=module,
                                                           totals_row=totals,
                                                           ws=ws,
                                                           idx_row=idx + 1)
                        break
                if not matched: 
                    logger.info(f"Found module {module_name_from_subplans} not existed in DUT, end parsing")
                    ws.append([idx+1, module_name_from_subplans,0,0,0,0,0,True,"0s" if self.time_unit == TimeUnit.HMS.value else 0])

        else: 
            for idx in range(len(modules_runner)):
                module=modules_runner[idx]
                assert module.get('name') is not None, "Can not get name of module"
                self.__parse_module_and_fill_excel(module=module,
                                                    totals_row=totals,
                                                    ws=ws,
                                                    idx_row=idx + 1)
        if self.time_unit == TimeUnit.S.value:
            totals[ReportColumn.EXECUTION_TIME.value] = str(totals[ReportColumn.EXECUTION_TIME.value]) + "s" 
        elif self.time_unit == TimeUnit.MS.value:
            totals[ReportColumn.EXECUTION_TIME.value] = str(totals[ReportColumn.EXECUTION_TIME.value]) + "ms" 

        total_row = ["Total", ""] + [totals[k] for k in headers[2:]]
        ws.append(total_row)

        output_file = f"./result/myresult_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        wb.save(output_file)
        logger.info(f"Saved result to {output_file}")

    def __parser_module(self,module : Element ) -> dict:
        name = module.get("name")
        logger.info(f"Parsing module {name}")

        execution_time = module.get("runtime")
        assert execution_time is not None, "Can not get runtime of module"
        execution_time = int(execution_time) 

        if self.time_unit == TimeUnit.HMS.value:
            execution_time = format_duration(math.ceil(execution_time/1000)) 
            logger.info(f"Execution Time: {execution_time}")
        elif self.time_unit == TimeUnit.S.value:
            execution_time = str(math.ceil(execution_time/1000)) + "s"
            logger.info(f"Execution Time: {execution_time} {self.time_unit}")
        elif self.time_unit == TimeUnit.MS.value:
            execution_time = str(execution_time) + "ms"
            logger.info(f"Execution Time: {execution_time} {self.time_unit}")
        
        assert module.get("done") is not None, "Can not get result of module"
        done = True if module.get("done") == "true" else False
        logger.info(f"Done: {done}")

        test_cases = module.findall(".TestCase")
        logger.info(f"Found {len(test_cases)} test cases")

        pass_count = 0
        fail_count = 0
        assumption_failure = 0
        ignored = 0
        for test_case in test_cases:
            logger.debug(f"Parsing test case {test_case.get('name')}")
            for test_step in test_case.findall(".Test"):
                test_step_name = test_step.get('name')
                test_step_result = test_step.get("result")
                if test_step_result == "fail":
                    fail_count += 1
                    logger.debug(f"Test case {test_step_name} failed")
                elif test_step_result == "ASSUMPTION_FAILURE":
                    assumption_failure += 1
                    logger.debug(f"Test case {test_step_name} assumption failure")
                elif test_step_result == "IGNORED":
                    ignored += 1
                    logger.debug(f"Test case {test_step_name} ignored")
                else:
                    pass_count += 1
                    logger.debug(f"Test case {test_step_name} passed")

        total_tests_str = module.get("total_tests")
        assert total_tests_str is not None, "Can not get total test of module"
        logger.info(f"Found {total_tests_str} tests")
        total_testcases = int(total_tests_str) if total_tests_str else 0

        assert total_testcases == pass_count + fail_count + assumption_failure + ignored, \
            f"Test count mismatch: {total_testcases} vs {pass_count} + {fail_count} + {assumption_failure} + {ignored}"
        
        return {
            ReportColumn.MODULES.value  : name,
            ReportColumn.PASSED.value   : pass_count,
            ReportColumn.FAILED.value   : fail_count,
            ReportColumn.ASSUMPTION_FAILURE.value : assumption_failure,
            ReportColumn.IGNORED.value  : ignored,
            ReportColumn.TOTAL.value    : total_testcases,
            ReportColumn.DONE.value     : done,
            ReportColumn.EXECUTION_TIME.value: execution_time
        }


    
